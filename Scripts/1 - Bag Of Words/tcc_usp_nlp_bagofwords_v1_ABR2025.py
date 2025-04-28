#Ambiente virtual
#/Users/maykow/Desktop/projects/empreducabot
#source empreducabot/bin/activate

from sentences import Sentence
from training_data_vABR2025_tcc_usp import TrainingData
import pandas as pd
from openpyxl import Workbook, load_workbook
import nltk


#para testes e acuracia
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm

CAMINHO = '/Users/maykow/desktop/projects/empreducabot/'

#este é o arquivo de frases adicionais sem numeros
arquivo = CAMINHO + "DadosTreinamentoPerguntasGerais.xlsx"

#este é o arquivo com palavras mais importantes repetidas para cada classe
#arquivo = CAMINHO + "Treinamento_TCC_USP_bagofwords_FAQ.xlsx"

TEST_FILE = "Testes_TCC_USP_perguntasgerais_bagofwords.xlsx"


def ler_excel_respostas(file, sheet):
    excel_data_df = pd.read_excel(file, sheet_name=sheet, index_col=None, usecols=['class', 'sentence'])
    return excel_data_df


def classificar_frase(frase):
    new_sentence = data.remove_accented_chars(frase)
    new_sentence2 = data.remove_stopwords_v2(new_sentence)
    classificacao, score, tabela_scores = data.calculate_score(new_sentence2)
    return classificacao


def verificar_classe(frase, classificacao_esperada):
    classificacao = classificar_frase(frase)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        #print(f"Teste OK para a frase: '{frase}' - Classificação: '{classificacao}'")
        #print('------------------------------------')
        return 'Classificação OK!'
    except AssertionError as e:
        #print(e)
        #print('------------------------------------')
        return classificacao


def executar_testes(arquivo):
    df2 = pd.read_excel(arquivo) #.read_excel lê a planilha
    wb = load_workbook(arquivo) #carrega o arquivo
    ws = wb.active #seleciona a primeira aba da planilha
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada)
        #print (f"Linha {index} calculada")
        ws.cell(row=index + 2, column=3, value=resultado) #+2 para ajustar o índice do Excel (cabeçalho)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)  # Soma 0 na coluna caso resultado esteja ok
        else:
            ws.cell(row=index + 2, column=4, value=1)  # Soma 0 na coluna caso resultado esteja ok
    wb.save(arquivo)

    #Soma dos erros
    soma = 0
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        valor = row[0]
        if isinstance(valor, (int, float)):
            soma += valor
    total_linhas = ws.max_row -1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print (f"Porcentagem de acerto: {(1-(soma/total_linhas))*100:.2f}%")


# Função para calcular acuracia, precisao, recall, f1
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true, y_pred, average='weighted', zero_division=0
    )
    return acc, precision, recall, f1

#iniciando bibliotecas
stopwords = nltk.corpus.stopwords.words('portuguese')
data = TrainingData(arquivo)
sentenca = Sentence()


# Testes de Usuário - menor amostra
# if __name__ == "__main__":
#
#     while (True):
#         resposta = input ("VOCE >>> ")
#         classificacao = classificar_frase(resposta)
#         print("CLASSIFICACAO >>>", classificacao)


if __name__ == "__main__":
    # DIRETORIO = os.getcwd()
    # arquivo = DIRETORIO + TEST_FILE

    executar_testes(TEST_FILE)

    # Ativa barra de progresso no apply
    tqdm.pandas()

    # Carregar planilha de teste
    df_test = pd.read_excel(TEST_FILE)

    # Aplicar o modelo
    df_test["class_pred"] = df_test["frase"].progress_apply(classificar_frase)

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])

    # Mostrar resultados
    print("Avaliação do modelo BagOfWords")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")