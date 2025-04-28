import os
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm
import unicodedata
import nltk

nltk.download('stopwords')
from nltk.corpus import stopwords

# Arquivos
#EXCEL_FILE = "planilha_perguntas_classificacao_longa.xlsx"
EXCEL_FILE = "Treinamento_TCC_USP_bagofwords_FAQ_tf_idf.xlsx"
if EXCEL_FILE == "planilha_perguntas_classificacao_longa.xlsx":
    VETORIZADOR_FILE = "vetorizador_tfidf.pkl"
    MODELO_FILE = "modelo_tfidf.pkl"
else:
    VETORIZADOR_FILE = "vetorizador_tfidf_amontoadopalavras.pkl"
    MODELO_FILE = "modelo_tfidf_amontoadopalavras.pkl"
TEST_FILE = "Testes_TCC_USP_perguntasgerais_tf_idf.xlsx"


# 📌 Pré-processamento
def remover_acentos(texto):
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8').lower()

def limpar_texto(texto):
    texto = remover_acentos(texto)
    palavras = texto.split()
    stop = set(stopwords.words('portuguese'))
    palavras = [p for p in palavras if p not in stop]
    return " ".join(palavras)

# 🧠 Treinamento inicial
df = pd.read_excel(EXCEL_FILE, sheet_name="trainning", usecols=["frase", "classificacao"])
df.dropna(inplace=True)
df["frase"] = df["frase"].astype(str).apply(limpar_texto)

vectorizer = TfidfVectorizer()
X = vectorizer.fit_transform(df["frase"])
y = df["classificacao"]

modelo = LogisticRegression(max_iter=1000)
modelo.fit(X, y)

# Função de classificação
def classificar_frase(frase_usuario):
    frase_limpa = limpar_texto(frase_usuario)
    vetor = vectorizer.transform([frase_limpa])
    pred = modelo.predict(vetor)[0]
    return pred

def verificar_classe(frase, classificacao_esperada):
    classificacao = classificar_frase(frase)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        return 'Classificação OK!'
    except AssertionError as e:
        return classificacao

def executar_testes(arquivo):
    df2 = pd.read_excel(arquivo)
    wb = load_workbook(arquivo)
    ws = wb.active
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada)
        ws.cell(row=index + 2, column=3, value=resultado)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)
        else:
            ws.cell(row=index + 2, column=4, value=1)
    wb.save(arquivo)

    soma = sum(1 for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True) if isinstance(row[0], (int, float)) and row[0] == 1)
    total_linhas = ws.max_row - 1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print(f"Porcentagem de acerto: {(1 - (soma / total_linhas)) * 100:.2f}%")


# Avaliação geral
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(y_true, y_pred, average='weighted', zero_division=0)
    return acc, precision, recall, f1


# Execução principal
if __name__ == "__main__":
    executar_testes(TEST_FILE)

    tqdm.pandas()
    df_test = pd.read_excel(TEST_FILE)
    df_test["class_pred"] = df_test["frase"].progress_apply(classificar_frase)

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])
    print("🎯 Avaliação do modelo TF-IDF")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")