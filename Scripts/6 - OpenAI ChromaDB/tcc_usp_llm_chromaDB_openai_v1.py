#RETRIEVAL
#pip install -r requirements.txt
#pip install httpx==0.27.2
import os
import pandas as pd

os.environ['OPENAI_API_KEY']= "OPENAI_KEY"


from langchain_community.document_loaders.pdf import PyPDFLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders.csv_loader import CSVLoader
#para testes e acuracia
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm


# caminho = 'planilha_perguntas_classificacao_longa.csv'
# TEST_FILE = "Testes_TCC_USP_perguntasgerais_openai_vectordb.xlsx"

caminho = 'Treinamento_TCC_USP_lancamentos_frasecompleta.csv'
TEST_FILE = "Testes_TCC_USP_lancamentos_versao_maior.xlsx"

loader=CSVLoader(caminho)
planilha = loader.load()


recur_split = RecursiveCharacterTextSplitter(
    chunk_size=500,
    chunk_overlap=50,
    separators=["\n\n", "\n", ".", " ", ""]
)

documents = recur_split.split_documents(planilha)

for i, doc in enumerate(documents):
    doc.metadata['doc_id'] = i

print(documents[2].metadata)

from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores.chroma import Chroma

embeddings_model = OpenAIEmbeddings()

#diretorio = 'arquivos/chroma_retrival_lucrefy_bd'
diretorio = 'arquivos/chroma_retrival_lucrefy_vendas_bd'

#COMENTAR NO CASO DA BASE JÁ ESTAR EM DISCO
vectordb = Chroma.from_documents(
    documents=documents,
    embedding=embeddings_model,
    persist_directory=diretorio
)


# vectordb = Chroma(
#     embedding_function=embeddings_model,
#     persist_directory=diretorio
# )


def classificacao_openai_chrome(pergunta, vectordb):
    doc = vectordb.max_marginal_relevance_search(pergunta, k=3, fetch_k=10)

    conteudo = doc[0].page_content  # conteúdo da célula
    linhas = conteudo.split("\n")

    for linha in linhas:
        if linha.lower().startswith("classificacao:"):
            return linha.split(":", 1)[1].strip()

    return "Classificação não encontrada"


def verificar_classe(frase, classificacao_esperada, vectordb):
    classificacao = classificacao_openai_chrome(frase, vectordb)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        #print(f"Teste OK para a frase: '{frase}' - Classificação: '{classificacao}'")
        #print('------------------------------------')
        return 'Classificação OK!'
    except AssertionError as e:
        #print(e)
        #print('------------------------------------')
        return classificacao


def executar_testes(arquivo, vectordb):
    df2 = pd.read_excel(arquivo) #.read_excel lê a planilha
    wb = load_workbook(arquivo) #carrega o arquivo
    ws = wb.active #seleciona a primeira aba da planilha
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada, vectordb)
        #print (f"Linha {index} calculada")
        ws.cell(row=index + 2, column=3, value=resultado) #+2 para ajustar o índice do Excel (cabeçalho)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)  # Soma 0 na coluna caso resultado esteja ok
        else:
            ws.cell(row=index + 2, column=4, value=1)  # Soma 0 na coluna caso resultado esteja ok
    wb.save(arquivo)

    #Soma dos erros
    soma = 0
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        valor = row[0]
        if isinstance(valor, (int, float)):
            soma += valor
    total_linhas = ws.max_row -1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print (f"Porcentagem de acerto: {(1-(soma/total_linhas))*100:.2f}%")


# Função para calcular acuracia, precisao, recall, f1
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true, y_pred, average='weighted', zero_division=0
    )
    return acc, precision, recall, f1

# if __name__=="__main__":
#     while True:
#         pergunta = input("Qual a sua pergunta?   >>>  ")
#         classificacao = classificacao_openai_chrome(pergunta, vectordb)
#         print(f"Classificação: {classificacao}")
#         print("------------------------............................-----------------")


# Interface simples para testes com volume maior localizado em excel já classificado
if __name__ == "__main__":
    # DIRETORIO = os.getcwd()
    # arquivo = DIRETORIO + TEST_FILE
    executar_testes(TEST_FILE, vectordb)

    # Ativa barra de progresso no apply
    tqdm.pandas()

    # Carregar planilha de teste
    df_test = pd.read_excel(TEST_FILE)

    # Aplicar o modelo
    df_test["class_pred"] = df_test["frase"].progress_apply(lambda frase: classificacao_openai_chrome(frase, vectordb))

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])

    # Mostrar resultados
    print("🎯 Avaliação do modelo OpenAI VectorDB")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")


