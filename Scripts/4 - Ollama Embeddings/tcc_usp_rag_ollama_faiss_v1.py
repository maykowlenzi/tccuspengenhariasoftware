import os
import pandas as pd
import numpy as np
import faiss
import requests

#para testes e acuracia
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm

# Arquivos persistentes
PLANILHA_ORIGINAL = "planilha_perguntas_classificacao_longa.xlsx"
INDEX_FILE = "faiss_index_ollama_longo.index"
METADADOS_FILE = "metadados_ollama_faiss_longo.csv"
TEST_FILE = "Testes_TCC_USP_perguntasgerais_ollama_embeddings.xlsx"

# === 1. Função para obter embeddings do Ollama (mxbai-embed-large)
def get_ollama_embedding(text, model="mxbai-embed-large"):
    url = "http://localhost:11434/api/embeddings"
    response = requests.post(url, json={"model": model, "prompt": text})
    response.raise_for_status()
    return response.json()["embedding"]

# === 2. Carrega ou cria embeddings e índice FAISS
def carregar_ou_criar_index():
    if os.path.exists(INDEX_FILE) and os.path.exists(METADADOS_FILE):
        print("Carregando índice e metadados do disco...")
        index = faiss.read_index(INDEX_FILE)
        df = pd.read_csv(METADADOS_FILE)
        df["embedding"] = df["embedding"].apply(eval)
    else:
        print("⚙Gerando embeddings com Ollama...")
        df = pd.read_excel(PLANILHA_ORIGINAL)
        df["embedding"] = df["frase"].apply(get_ollama_embedding)

        # Salva metadados
        df.to_csv(METADADOS_FILE, index=False)

        # Cria índice FAISS
        dim = len(df["embedding"][0])
        index = faiss.IndexFlatL2(dim)
        index.add(np.array(df["embedding"].tolist()).astype("float32"))
        faiss.write_index(index, INDEX_FILE)

    return df, index

# === 3. Busca os k mais parecidos
def buscar_contexto(frase_usuario, df, index, k=1):
    emb = np.array(get_ollama_embedding(frase_usuario)).astype("float32").reshape(1, -1)
    D, I = index.search(emb, k)
    #print("indice >>> ", I)
    exemplos = []
    for idx in I[0]:
        exemplo = f"Frase: {df.iloc[idx]['frase']}\nClassificação: {df.iloc[idx]['classificacao']}"
        exemplos.append(exemplo)
    #return "\n\n".join(exemplos)
    #return exemplos[0]
    #return I[0]
    #for indice in range (0, k):
    #    print("Classificacao ", indice+1, df.iloc[I[0][indice]]['classificacao'], ">>>>", df.iloc[I[0][indice]]['frase'])
    resultado = df.iloc[I[0][0]]
    return resultado['classificacao']

# === 4. Gera resposta com modelo local (ex: llama3.2)
def classificar_com_ollama(frase_usuario, df, index, model="llama3.2"):
    contexto = buscar_contexto(frase_usuario, df, index)
#
#     prompt = f"""
# Você é um assistente de classificação de frases financeiras.
#
# Com base nos exemplos abaixo, classifique corretamente a nova frase de um cliente.
#
# ### Exemplos
# {contexto}
#
# ### Nova frase
# Frase: {frase_usuario}
# Classificação:"""
#
#     url = "http://localhost:11434/api/generate"
#     response = requests.post(url, json={
#         "model": model,
#         "prompt": prompt,
#         "temperature": 0.0,
#         "stream": False
#     })
#     response.raise_for_status()
#     return response.json()["response"].strip()

    return contexto


def verificar_classe(frase, classificacao_esperada, df_teste, index):
    classificacao = buscar_contexto(frase, df_teste, index)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        #print(f"Teste OK para a frase: '{frase}' - Classificação: '{classificacao}'")
        #print('------------------------------------')
        return 'Classificação OK!'
    except AssertionError as e:
        #print(e)
        #print('------------------------------------')
        return classificacao


def executar_testes(arquivo, df_teste, index_teste):
    df2 = pd.read_excel(arquivo) #.read_excel lê a planilha
    wb = load_workbook(arquivo) #carrega o arquivo
    ws = wb.active #seleciona a primeira aba da planilha
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada, df_teste, index_teste)
        #print (f"Linha {index} calculada")
        ws.cell(row=index + 2, column=3, value=resultado) #+2 para ajustar o índice do Excel (cabeçalho)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)  # Soma 0 na coluna caso resultado esteja ok
        else:
            ws.cell(row=index + 2, column=4, value=1)  # Soma 0 na coluna caso resultado esteja ok
    wb.save(arquivo)

    #Soma dos erros
    soma = 0
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        valor = row[0]
        if isinstance(valor, (int, float)):
            soma += valor
    total_linhas = ws.max_row -1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print (f"Porcentagem de acerto: {(1-(soma/total_linhas))*100:.2f}%")


# Função para calcular acuracia, precisao, recall, f1
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true, y_pred, average='weighted', zero_division=0
    )
    return acc, precision, recall, f1

# === Execução Testes Unitários
# if __name__ == "__main__":
#     df, index = carregar_ou_criar_index()
#
#     while True:
#         nova_frase = input("PERGUNTA >>> ")
#         #resultado = classificar_com_ollama(nova_frase, df, index)
#         resultado = buscar_contexto(nova_frase, df, index)
#         print(f" Classificação sugerida: {resultado}")


# Interface simples para testes com volume maior localizado em excel já classificado
if __name__ == "__main__":
    # DIRETORIO = os.getcwd()
    # arquivo = DIRETORIO + TEST_FILE
    df, index = carregar_ou_criar_index()
    #executar_testes(TEST_FILE, df, index)

    # Ativa barra de progresso no apply
    tqdm.pandas()

    # Carregar planilha de teste
    df_test = pd.read_excel(TEST_FILE)

    # Aplicar o modelo
    df_test["class_pred"] = df_test["frase"].progress_apply(lambda frase: buscar_contexto(frase, df, index))

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])

    # Mostrar resultados
    print("🎯 Avaliação do modelo Ollama Embeddings")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")