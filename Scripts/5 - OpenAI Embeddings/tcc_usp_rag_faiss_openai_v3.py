import os
import pandas as pd
import faiss
#para fazer indexação por similaridade
import openai
import numpy as np

#para testes e acuracia
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm

# Sua chave da OpenAI
openai.api_key = "OPENAI_API"

# Nome dos arquivos persistentes
INDEX_FILE = "faiss_classificacao_longa_openai.index"
METADADOS_FILE = "metadados_classificacao_longa_openai.csv"
PLANILHA_ORIGINAL = "planilha_perguntas_classificacao_longa.xlsx"
TEST_FILE = "Testes_TCC_USP_perguntasgerais_openai_embeddings.xlsx"

# Função para gerar embeddings da OpenAI
def get_embedding(text, model="text-embedding-3-small"):
    response = openai.embeddings.create(input=[text], model=model)
    return response.data[0].embedding

# Carregamento com persistência
if os.path.exists(INDEX_FILE) and os.path.exists(METADADOS_FILE):
    # Carrega índice FAISS e dados
    index = faiss.read_index(INDEX_FILE)
    df = pd.read_csv(METADADOS_FILE)
else:
    # ⚙Cria tudo do zero
    df = pd.read_excel(PLANILHA_ORIGINAL)

    # Validação
    assert 'frase' in df.columns and 'classificacao' in df.columns

    # Gera embeddings
    df['embedding'] = df['frase'].apply(get_embedding)

    # Cria índice FAISS
    dimension = len(df['embedding'].iloc[0])
    index = faiss.IndexFlatL2(dimension)
    embedding_matrix = np.array(df['embedding'].tolist()).astype('float32')
    index.add(embedding_matrix)

    # Salva para reuso
    faiss.write_index(index, INDEX_FILE)
    df[['frase', 'classificacao']].to_csv(METADADOS_FILE, index=False)

# Função de busca por similaridade
def classificar_frase(nova_frase, k=1):
    nova_embedding = np.array([get_embedding(nova_frase)]).astype('float32')
    distancias, indices = index.search(nova_embedding, k)
    resultado = df.iloc[indices[0][0]]
    return resultado['classificacao']


def verificar_classe(frase, classificacao_esperada):
    classificacao = classificar_frase(frase)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        #print(f"Teste OK para a frase: '{frase}' - Classificação: '{classificacao}'")
        #print('------------------------------------')
        return 'Classificação OK!'
    except AssertionError as e:
        #print(e)
        #print('------------------------------------')
        return classificacao


def executar_testes(arquivo):
    df2 = pd.read_excel(arquivo) #.read_excel lê a planilha
    wb = load_workbook(arquivo) #carrega o arquivo
    ws = wb.active #seleciona a primeira aba da planilha
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada)
        #print (f"Linha {index} calculada")
        ws.cell(row=index + 2, column=3, value=resultado) #+2 para ajustar o índice do Excel (cabeçalho)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)  # Soma 0 na coluna caso resultado esteja ok
        else:
            ws.cell(row=index + 2, column=4, value=1)  # Soma 0 na coluna caso resultado esteja ok
    wb.save(arquivo)

    #Soma dos erros
    soma = 0
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        valor = row[0]
        if isinstance(valor, (int, float)):
            soma += valor
    total_linhas = ws.max_row -1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print (f"Porcentagem de acerto: {(1-(soma/total_linhas))*100:.2f}%")


# Função para calcular acuracia, precisao, recall, f1
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true, y_pred, average='weighted', zero_division=0
    )
    return acc, precision, recall, f1


# Exemplo de uso interativo
# if __name__ == "__main__":
#     while True:
#         entrada = input("\nPERGUNTA >>> ")
#         if entrada.lower() == "sair":
#             break
#         try:
#             classificacao = classificar_frase(entrada)
#             print(f"Classificação mais provável: {classificacao}")
#         except Exception as e:
#             print(f"Erro: {e}")



# Interface simples para testes com volume maior localizado em excel já classificado
if __name__ == "__main__":
    # DIRETORIO = os.getcwd()
    # arquivo = DIRETORIO + TEST_FILE
    executar_testes(TEST_FILE)

    # Ativa barra de progresso no apply
    tqdm.pandas()

    # Carregar planilha de teste
    df_test = pd.read_excel(TEST_FILE)

    # Aplicar o modelo
    df_test["class_pred"] = df_test["frase"].progress_apply(classificar_frase)

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])

    # Mostrar resultados
    print("🎯 Avaliação do modelo OpenAI Embeddings")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")
