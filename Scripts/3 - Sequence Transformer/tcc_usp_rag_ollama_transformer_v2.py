import os
import pandas as pd
import numpy as np
import faiss
from sentence_transformers import SentenceTransformer
#gera embeddings ou vetores de frases sem OpenAI
import requests
#para testes e acuracia
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from tqdm import tqdm


# Arquivos
EXCEL_FILE = "planilha_perguntas_classificacao_longa.xlsx"
#EXCEL_FILE = "Treinamento_TCC_USP_bagofwords_FAQ.xlsx"
if EXCEL_FILE == "planilha_perguntas_classificacao_longa.xlsx":
    INDEX_FILE = "faiss_classificacao_longa_transformer.index"
    METADADOS_FILE = "metadados_classificacao_longa_transformer.csv"
else:
    INDEX_FILE = "faiss_classificacao_longa_transformer_amontoadopalavras.index"
    METADADOS_FILE = "metadados_classificacao_longa_transformer_amontoadopalavras.csv"
TEST_FILE = "Testes_TCC_USP_perguntasgerais_ollama.xlsx"


# Modelo local (usa CPU por padrão)
model = SentenceTransformer("all-MiniLM-L6-v2")
# pequeno e rápido, ótimo para RAG. Transforma texto em vetores. Roda localmente e tem bom resultado para similaridade

# Função para gerar embeddings. Recebe texto e retorna um vetor
def get_embedding(text):
    return model.encode([text])[0]  # retorna array 1D

# Carrega dados com persistência
if os.path.exists(INDEX_FILE) and os.path.exists(METADADOS_FILE):
    #aqui, se já existir os arquivos de index e csv, -> persistência dos dados
    index = faiss.read_index(INDEX_FILE)
    df = pd.read_csv(METADADOS_FILE)
else:
    df = pd.read_excel(EXCEL_FILE)
    assert 'frase' in df.columns and 'classificacao' in df.columns

    # Gera embeddings
    df['embedding'] = df['frase'].apply(get_embedding)
    #aqui pega o texto e gera um vetor

    # Cria índice FAISS
    dimension = len(df['embedding'].iloc[0])
    index = faiss.IndexFlatL2(dimension) #baseada em distância L2
    embeddings_matrix = np.stack(df['embedding'].values).astype('float32')
    index.add(embeddings_matrix)

    # Salva FAISS + metadados
    faiss.write_index(index, INDEX_FILE)
    df[['frase', 'classificacao']].to_csv(METADADOS_FILE, index=False)

# Função de classificação
def classificar_frase(frase_usuario, k=1):
    #k é o indice de k frases mais próximas -> retorna a melhor classificacao de acordo com o treinamento
    embedding = np.array([get_embedding(frase_usuario)]).astype('float32')
    distancias, indices = index.search(embedding, k)
    resultado = df.iloc[indices[0][0]]
    # loop para checar a qualidade de respostas
    for i in range(0, k-1):
        teste = df.iloc[indices[0][i]]
        print(f"resultado {i}: ", teste['classificacao'])

    resposta = resultado['classificacao']

    return resposta


def responder_com_ollama(frase_usuario, classificacao):
    prompt = f"""
A frase: "{frase_usuario}" foi classificada como: "{classificacao}".

Explique de forma clara e simples por que essa classificação faz sentido, como se estivesse ajudando um pequeno empreendedor a entender.
"""
    try:
        response = requests.post(
            "http://localhost:11434/api/generate",
            json={
                "model": "llama3.2",
                "prompt": prompt,
                "stream": False
            }
        )
        if response.status_code == 200:
            return response.json()["response"]
        else:
            return f"[Erro no Ollama] Código: {response.status_code}"
    except Exception as e:
        return f"[Erro de conexão com Ollama] {str(e)}"


def verificar_classe(frase, classificacao_esperada):
    classificacao = classificar_frase(frase)
    try:
        assert classificacao == classificacao_esperada, f"Erro: {classificacao} != {classificacao_esperada} para a frase: '{frase}'"
        #print(f"Teste OK para a frase: '{frase}' - Classificação: '{classificacao}'")
        #print('------------------------------------')
        return 'Classificação OK!'
    except AssertionError as e:
        #print(e)
        #print('------------------------------------')
        return classificacao


def executar_testes(arquivo):
    df2 = pd.read_excel(arquivo) #.read_excel lê a planilha
    wb = load_workbook(arquivo) #carrega o arquivo
    ws = wb.active #seleciona a primeira aba da planilha
    for index, row in tqdm(df2.iterrows(), total=len(df2), desc="Testando frases"):
        frase = row['frase']
        classificacao_esperada = row['class']
        resultado = verificar_classe(frase, classificacao_esperada)
        #print (f"Linha {index} calculada")
        ws.cell(row=index + 2, column=3, value=resultado) #+2 para ajustar o índice do Excel (cabeçalho)
        if resultado == 'Classificação OK!':
            ws.cell(row=index + 2, column=4, value=0)  # Soma 0 na coluna caso resultado esteja ok
        else:
            ws.cell(row=index + 2, column=4, value=1)  # Soma 0 na coluna caso resultado esteja ok
    wb.save(arquivo)

    #Soma dos erros
    soma = 0
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        valor = row[0]
        if isinstance(valor, (int, float)):
            soma += valor
    total_linhas = ws.max_row -1
    print(f"Total de Linhas {total_linhas}")
    print(f"Total de Linhas que não bateram classificação {soma}")
    print("------------------------------")
    print (f"Porcentagem de acerto: {(1-(soma/total_linhas))*100:.2f}%")


# Função para calcular acuracia, precisao, recall, f1
def avaliar(y_true, y_pred):
    acc = accuracy_score(y_true, y_pred)
    precision, recall, f1, _ = precision_recall_fscore_support(
        y_true, y_pred, average='weighted', zero_division=0
    )
    return acc, precision, recall, f1

# Interface simples para testes unitários
# if __name__ == "__main__":
#     while True:
#         frase = input("\nPERGUNTA : ")
#         if frase.lower() == "sair":
#             break
#         try:
#             classificacao = classificar_frase(frase)
#             print(f"Classificação mais provável: {classificacao}")
#             #explicacao = responder_com_ollama(frase, classificacao)
#             #print(f"\nExplicação do modelo (Ollama):\n{explicacao}")
#         except Exception as e:
#             print(f"Erro: {e}")

# Interface simples para testes com volume maior localizado em excel já classificado
if __name__ == "__main__":
    # DIRETORIO = os.getcwd()
    # arquivo = DIRETORIO + TEST_FILE
    executar_testes(TEST_FILE)

    # Ativa barra de progresso no apply
    tqdm.pandas()

    # Carregar planilha de teste
    df_test = pd.read_excel(TEST_FILE)

    # Aplicar o modelo
    df_test["class_pred"] = df_test["frase"].progress_apply(classificar_frase)

    acc, prec, rec, f1 = avaliar(df_test["class"], df_test["class_pred"])

    # Mostrar resultados
    print("🎯 Avaliação do modelo SentenceTransformer")
    print(f"Acurácia : {acc:.3f}")
    print(f"Precisão : {prec:.3f}")
    print(f"Recall   : {rec:.3f}")
    print(f"F1-score : {f1:.3f}")